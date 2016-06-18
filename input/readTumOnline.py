#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division

import sys
import os
PATHS = os.getcwd().split('/')
PROJECT_PATH = ''
for p in PATHS:
    PROJECT_PATH += '%s' % p
    if p == 'examination-scheduling':
        break
sys.path.append(PROJECT_PATH)

from time import time
import random 
import openpyxl

def read_real_data():
    wb = openpyxl.load_workbook('%s\input\Data\TumOnline\Read15S.xlsx' % (PROJECT_PATH))
    ws = wb.get_sheet_by_name('Conflicts')

    maxRow = ws.max_row

    # create number of exams
    n = maxRow-1

    # create Conflict Matrix
    Q = [[int(cell.value) if not cell.value is None else 0 for cell in row]   for row in ws.iter_rows("B2:%s%s" % (openpyxl.utils.get_column_letter(maxRow),maxRow))]


    
    return 


if __name__ == '__main__':
    read_real_data()